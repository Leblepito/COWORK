"""
COWORK.ARMY v7.0 — Cargo File Extractor
Extracts readable text from any file format for routing analysis.

Supported formats:
- Text: .txt, .md, .log, .py, .js, .ts, .html, .css, .json, .xml, .yaml, .yml, .sql
- Spreadsheet: .csv, .xlsx, .xls
- Document: .pdf (metadata), .docx (python-docx if available)
- Image: returns metadata only
- Binary/Unknown: returns filename + size metadata
"""
import io
import csv
import json
from pathlib import Path

MAX_TEXT_LENGTH = 50_000  # chars


def extract_text(raw: bytes, filename: str = "", content_type: str = "") -> str:
    """
    Extract readable text from raw file bytes.
    Never raises — always returns a string (even if just metadata).
    """
    ext = Path(filename).suffix.lower() if filename else ""
    ct = (content_type or "").lower()

    # ── Plain text formats ──────────────────────────────────────────────────
    text_exts = {".txt", ".md", ".log", ".py", ".js", ".ts", ".jsx", ".tsx",
                 ".html", ".htm", ".css", ".json", ".xml", ".yaml", ".yml",
                 ".sql", ".sh", ".env", ".toml", ".ini", ".cfg"}
    text_cts = {"text/", "application/json", "application/xml",
                "application/javascript", "application/x-yaml"}

    is_text_ext = ext in text_exts
    is_text_ct = any(ct.startswith(p) for p in text_cts)

    if is_text_ext or is_text_ct:
        try:
            return raw.decode("utf-8", errors="replace")[:MAX_TEXT_LENGTH]
        except Exception:
            pass

    # ── CSV ────────────────────────────────────────────────────────────────
    if ext == ".csv" or ct == "text/csv":
        try:
            text = raw.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = []
            for i, row in enumerate(reader):
                if i > 500:
                    break
                rows.append(", ".join(row))
            return "\n".join(rows)[:MAX_TEXT_LENGTH]
        except Exception:
            try:
                return raw.decode("utf-8", errors="replace")[:MAX_TEXT_LENGTH]
            except Exception:
                pass

    # ── Excel .xlsx ────────────────────────────────────────────────────────
    if ext in (".xlsx", ".xlsm") or "spreadsheetml" in ct:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parts.append(f"[Sheet: {sheet_name}]")
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    if row_count > 1000:
                        break
                    row_str = "\t".join(str(v) if v is not None else "" for v in row)
                    if row_str.strip():
                        parts.append(row_str)
                        row_count += 1
            return "\n".join(parts)[:MAX_TEXT_LENGTH]
        except Exception as e:
            return f"[Excel dosyası: {filename}, boyut: {len(raw)} byte, okuma hatası: {e}]"

    # ── Excel .xls (eski format) ───────────────────────────────────────────
    if ext == ".xls" or ct == "application/vnd.ms-excel":
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=raw)
            parts = []
            for sheet in wb.sheets():
                parts.append(f"[Sheet: {sheet.name}]")
                for rx in range(min(sheet.nrows, 1000)):
                    row_vals = [str(sheet.cell_value(rx, cx)) for cx in range(sheet.ncols)]
                    parts.append("\t".join(row_vals))
            return "\n".join(parts)[:MAX_TEXT_LENGTH]
        except Exception:
            # xlrd başarısız olursa openpyxl ile dene
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                parts = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    parts.append(f"[Sheet: {sheet_name}]")
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i > 1000:
                            break
                        parts.append("\t".join(str(v) if v is not None else "" for v in row))
                return "\n".join(parts)[:MAX_TEXT_LENGTH]
            except Exception as e2:
                return f"[XLS dosyası: {filename}, boyut: {len(raw)} byte, okuma hatası: {e2}]"

    # ── PDF ────────────────────────────────────────────────────────────────
    if ext == ".pdf" or ct == "application/pdf":
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                pages = []
                for i, page in enumerate(pdf.pages):
                    if i > 20:
                        break
                    text = page.extract_text() or ""
                    pages.append(text)
                return "\n".join(pages)[:MAX_TEXT_LENGTH]
        except Exception:
            pass
        # Fallback: try basic text extraction
        try:
            text = raw.decode("latin-1", errors="replace")
            # Extract printable ASCII runs
            import re
            runs = re.findall(r'[ -~]{4,}', text)
            return " ".join(runs)[:MAX_TEXT_LENGTH]
        except Exception:
            return f"[PDF dosyası: {filename}, boyut: {len(raw)} byte]"

    # ── Word .docx ────────────────────────────────────────────────────────
    if ext == ".docx" or "wordprocessingml" in ct:
        try:
            import docx
            doc = docx.Document(io.BytesIO(raw))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paragraphs)[:MAX_TEXT_LENGTH]
        except Exception:
            # Fallback: extract text from XML
            try:
                import zipfile, re
                with zipfile.ZipFile(io.BytesIO(raw)) as z:
                    xml = z.read("word/document.xml").decode("utf-8", errors="replace")
                    text = re.sub(r"<[^>]+>", " ", xml)
                    text = re.sub(r"\s+", " ", text).strip()
                    return text[:MAX_TEXT_LENGTH]
            except Exception as e:
                return f"[Word dosyası: {filename}, boyut: {len(raw)} byte, okuma hatası: {e}]"

    # ── Image formats ─────────────────────────────────────────────────────
    image_exts = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".svg",
                  ".tiff", ".tif", ".ico"}
    if ext in image_exts or ct.startswith("image/"):
        return f"[Görsel dosyası: {filename}, format: {ext or ct}, boyut: {len(raw)} byte]"

    # ── ZIP / Archive ─────────────────────────────────────────────────────
    if ext in (".zip", ".tar", ".gz", ".rar", ".7z") or ct in ("application/zip", "application/x-tar"):
        try:
            import zipfile
            with zipfile.ZipFile(io.BytesIO(raw)) as z:
                names = z.namelist()
                return f"[ZIP arşivi: {filename}, içerik: {', '.join(names[:20])}]"
        except Exception:
            return f"[Arşiv dosyası: {filename}, boyut: {len(raw)} byte]"

    # ── Generic binary fallback ───────────────────────────────────────────
    # Try UTF-8 decode first
    try:
        decoded = raw.decode("utf-8", errors="strict")
        return decoded[:MAX_TEXT_LENGTH]
    except UnicodeDecodeError:
        pass

    # Try latin-1 and extract printable runs
    try:
        import re
        text = raw.decode("latin-1", errors="replace")
        runs = re.findall(r'[ -~\n\t]{4,}', text)
        if runs:
            extracted = " ".join(runs)[:MAX_TEXT_LENGTH]
            if len(extracted) > 50:
                return extracted
    except Exception:
        pass

    return f"[Binary dosyası: {filename}, format: {ext or ct or 'bilinmiyor'}, boyut: {len(raw)} byte]"
